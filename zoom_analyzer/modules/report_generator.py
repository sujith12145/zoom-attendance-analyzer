"""
Module: report_generator.py
=============================
Generate downloadable reports in:
  • CSV   – simple pandas export
  • Excel – multi-sheet workbook with formatting (openpyxl / xlsxwriter)
  • PDF   – formatted report with tables and summary (fpdf2)
"""

import io
import os
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Optional

# Excel
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# PDF
try:
    from fpdf import FPDF
    FPDF_OK = True
except ImportError:
    FPDF_OK = False

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import APP_NAME, COLOR_GREEN, COLOR_YELLOW, COLOR_RED


# ─────────────────────────────────────────────────────────────────────────────
# CSV
# ─────────────────────────────────────────────────────────────────────────────

def generate_csv(df: pd.DataFrame) -> bytes:
    """Return UTF-8 CSV bytes."""
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode('utf-8')


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

def _hex_to_argb(hex_color: str) -> str:
    """Convert #RRGGBB to ARGB string."""
    return 'FF' + hex_color.lstrip('#').upper()


def generate_excel(
    attendance_df: pd.DataFrame,
    cumulative_df: Optional[pd.DataFrame] = None,
    history_df: Optional[pd.DataFrame] = None,
    title: str = "Attendance Report",
) -> bytes:
    """Return Excel workbook bytes with multiple sheets and formatting."""
    if not OPENPYXL_OK:
        # Fallback: plain xlsx via pandas
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            attendance_df.to_excel(writer, index=False, sheet_name='Attendance')
        return buf.getvalue()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Attendance ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Attendance'
    _write_header(ws1, f"{title} – {datetime.now().strftime('%Y-%m-%d')}")
    _df_to_sheet(ws1, attendance_df, start_row=3, status_col='status')

    # ── Sheet 2: Cumulative Summary ─────────────────────────────────────────
    if cumulative_df is not None and not cumulative_df.empty:
        ws2 = wb.create_sheet('Cumulative Summary')
        _write_header(ws2, 'Cumulative Attendance Summary')
        _df_to_sheet(ws2, cumulative_df, start_row=3, status_col=None, pct_col='attendance_pct')

    # ── Sheet 3: Full History ───────────────────────────────────────────────
    if history_df is not None and not history_df.empty:
        ws3 = wb.create_sheet('Full History')
        _write_header(ws3, 'Complete Attendance History')
        _df_to_sheet(ws3, history_df, start_row=3, status_col='status')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws, title: str):
    ws['A1'] = title
    ws['A1'].font      = Font(bold=True, size=14, color='FFFFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor='FF2C3E50')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A1:{get_column_letter(10)}1')


def _df_to_sheet(ws, df: pd.DataFrame, start_row: int,
                 status_col: Optional[str], pct_col: Optional[str] = None):
    """Write DataFrame to worksheet with formatting."""
    if df.empty:
        return

    header_fill  = PatternFill('solid', fgColor='FF3498DB')
    header_font  = Font(bold=True, color='FFFFFFFF')
    alt_fill     = PatternFill('solid', fgColor='FFF0F4F8')
    border_side  = Side(style='thin', color='FFBDC3C7')
    thin_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)

    # Header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name).replace('_', ' ').title())
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border    = thin_border

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, (col_name, val) in enumerate(row.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            cell.border    = thin_border

            # Status colour
            if status_col and col_name == status_col:
                if str(val) == 'Present':
                    cell.fill = PatternFill('solid', fgColor=_hex_to_argb(COLOR_GREEN))
                    cell.font = Font(color='FFFFFFFF', bold=True)
                elif str(val) == 'Absent':
                    cell.fill = PatternFill('solid', fgColor=_hex_to_argb(COLOR_RED))
                    cell.font = Font(color='FFFFFFFF', bold=True)

            # Percentage colour
            if pct_col and col_name == pct_col:
                try:
                    pct = float(val)
                    if pct >= 75:
                        cell.fill = PatternFill('solid', fgColor=_hex_to_argb(COLOR_GREEN))
                        cell.font = Font(color='FFFFFFFF')
                    elif pct >= 50:
                        cell.fill = PatternFill('solid', fgColor=_hex_to_argb(COLOR_YELLOW))
                    else:
                        cell.fill = PatternFill('solid', fgColor=_hex_to_argb(COLOR_RED))
                        cell.font = Font(color='FFFFFFFF')
                except (ValueError, TypeError):
                    pass

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────

def generate_pdf(
    attendance_df: pd.DataFrame,
    cumulative_df: Optional[pd.DataFrame] = None,
    class_name: str = "Class",
    class_date: str = "",
    summary: Optional[dict] = None,
) -> bytes:
    """Return PDF bytes with a formatted attendance report."""
    if not FPDF_OK:
        raise ImportError("fpdf2 is not installed. Run: pip install fpdf2")

    pdf = _AttendancePDF(class_name=class_name, class_date=class_date)
    pdf.add_page()

    # ── Summary block ──────────────────────────────────────────────────────
    if summary:
        pdf.section_title("Class Summary")
        for k, v in summary.items():
            pdf.key_value(str(k).replace('_', ' ').title(), str(v))
        pdf.ln(4)

    # ── Attendance table ────────────────────────────────────────────────────
    display_cols = ['roll_number', 'name', 'total_minutes', 'sessions', 'status']
    display_cols = [c for c in display_cols if c in attendance_df.columns]
    pdf.section_title("Attendance Record")
    pdf.attendance_table(attendance_df[display_cols])

    # ── Cumulative summary ──────────────────────────────────────────────────
    if cumulative_df is not None and not cumulative_df.empty:
        pdf.add_page()
        cum_cols = ['roll_number', 'name', 'total_classes', 'present_count',
                    'attendance_pct', 'risk_level']
        cum_cols = [c for c in cum_cols if c in cumulative_df.columns]
        pdf.section_title("Cumulative Attendance Summary")
        pdf.attendance_table(cumulative_df[cum_cols], pct_col='attendance_pct')

    return bytes(pdf.output())


class _AttendancePDF(FPDF):
    def __init__(self, class_name: str, class_date: str):
        super().__init__()
        self.class_name = class_name
        self.class_date = class_date
        self.set_auto_page_break(auto=True, margin=15)
        self.set_margins(15, 15, 15)

    def header(self):
        self.set_fill_color(44, 62, 80)
        self.rect(0, 0, 210, 18, 'F')
        self.set_text_color(255, 255, 255)
        self.set_font('Helvetica', 'B', 13)
        self.set_y(4)
        self.cell(0, 10, f'{APP_NAME}', align='C')
        self.set_y(18)
        self.set_text_color(0)

    def footer(self):
        self.set_y(-12)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128)
        self.cell(0, 8, f'Page {self.page_no()} | Generated {datetime.now().strftime("%Y-%m-%d %H:%M")}', align='C')

    def section_title(self, title: str):
        self.set_font('Helvetica', 'B', 11)
        self.set_fill_color(52, 152, 219)
        self.set_text_color(255)
        self.cell(0, 8, f'  {title}', fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0)
        self.ln(2)

    def key_value(self, key: str, val: str):
        self.set_font('Helvetica', 'B', 9)
        self.cell(60, 6, key + ':')
        self.set_font('Helvetica', '', 9)
        self.cell(0, 6, val, new_x="LMARGIN", new_y="NEXT")

    def attendance_table(self, df: pd.DataFrame, pct_col: Optional[str] = None):
        if df.empty:
            self.set_font('Helvetica', 'I', 9)
            self.cell(0, 6, 'No data available.', new_x="LMARGIN", new_y="NEXT")
            return

        col_widths = self._calc_col_widths(df)
        available  = 180

        # Scale widths
        total_w = sum(col_widths.values())
        if total_w > available:
            scale = available / total_w
            col_widths = {k: v * scale for k, v in col_widths.items()}

        # Header
        self.set_font('Helvetica', 'B', 8)
        self.set_fill_color(52, 73, 94)
        self.set_text_color(255)
        for col in df.columns:
            self.cell(col_widths[col], 7, str(col).replace('_', ' ').title(),
                      border=1, fill=True, align='C')
        self.ln()
        self.set_text_color(0)

        # Data rows
        self.set_font('Helvetica', '', 7.5)
        for i, (_, row) in enumerate(df.iterrows()):
            if i % 2 == 0:
                self.set_fill_color(245, 248, 250)
            else:
                self.set_fill_color(255, 255, 255)

            for col in df.columns:
                val = str(row[col]) if pd.notna(row[col]) else ''
                fill_flag = i % 2 == 0

                # Status cell colour
                if col == 'status':
                    if val == 'Present':
                        self.set_fill_color(46, 204, 113)
                        self.set_text_color(255)
                        fill_flag = True
                    elif val == 'Absent':
                        self.set_fill_color(231, 76, 60)
                        self.set_text_color(255)
                        fill_flag = True
                    else:
                        self.set_fill_color(243, 156, 18)
                        fill_flag = True

                # Percentage colour
                if pct_col and col == pct_col:
                    try:
                        pct = float(val)
                        if pct >= 75:
                            self.set_fill_color(46, 204, 113)
                            self.set_text_color(255)
                        elif pct >= 50:
                            self.set_fill_color(243, 156, 18)
                        else:
                            self.set_fill_color(231, 76, 60)
                            self.set_text_color(255)
                        fill_flag = True
                    except (ValueError, TypeError):
                        pass

                self.cell(col_widths[col], 6, val[:28],
                          border=1, fill=fill_flag, align='C')
                self.set_text_color(0)

            self.ln()

    def _calc_col_widths(self, df: pd.DataFrame) -> dict:
        """Rough column widths based on content."""
        widths = {}
        for col in df.columns:
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df.empty else 0,
            )
            widths[col] = min(max_len * 2.2 + 4, 55)
        return widths
